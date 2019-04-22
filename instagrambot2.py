#!/usr/bin/env python
# -*- coding: utf-8 -*-

import requests
import xlrd
import openpyxl
from time import sleep
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver import ActionChains
from datetime import datetime

erro_dropdown = 0

#a function to check if the comment was answered
def checkAnswer(name, listOfNames, listOfMessages):
	counterName = 0

	while counterName < len(listOfNames):
		if listOfNames[counterName] == "DEPILACAO.LASER.BR":
			if listOfMessages[counterName].find(name) is not -1:
				return True

		counterName = counterName + 1

	return False

#get the name and the message
def getNameAndMessage(browser):
	auxCounter = 0

	auxText = []
	auxName = []

	try:

		while auxCounter < len(browser.find_elements_by_xpath("//ul[@class='uiList _4kg  _4kt _6-h _6-j']/div")):
			auxText.append(browser.find_element_by_xpath("//ul[@class='uiList _4kg  _4kt _6-h _6-j']/div[" + str(auxCounter+1) + "]/div/div/div[2]/div/span/span[2]").text.upper())
			auxName.append(browser.find_element_by_xpath("//ul[@class='uiList _4kg  _4kt _6-h _6-j']/div[" + str(auxCounter+1) + "]/div/div/div[2]/div/span/span[1]").text.upper())
			print auxName[auxCounter]
			auxCounter = auxCounter + 1

	except:
		pass

	return auxName, auxText

#answer the messages in the screen
def answer(browser, auxName, auxText):
	auxCounter = 0

	respostas = 0

	global_counter = len(auxText)

	print global_counter

	while auxCounter < global_counter:
		try:
			if auxText[auxCounter].find("QUERO") is not -1:
				print "chegou"

				if (not checkAnswer(auxName[auxCounter], auxName, auxText)) and (auxName[auxCounter] != "DEPILACAO.LASER.BR"):
					#click to answer
					elem = browser.find_element_by_xpath("//ul[@class='uiList _4kg  _4kt _6-h _6-j']/div[" + str(auxCounter+1+respostas) + "]/div/div/div[3]/div[1]/a[2]")
					browser.execute_script("arguments[0].click();", elem)
					sleep(1)
					browser.find_element_by_xpath("//div[@class='_4u-c _5n4j']/div[1]/div[2]/div/div/div/div/label/input").send_keys(u"Inscreva-se clicando no link abaixo da imagem, aguarde contato telefônico e aproveite.\n")
					sleep(3)
					respostas = respostas + 1

			if (auxText[auxCounter].find("ONDE") is not -1) or (auxText[auxCounter].find(u"ENDEREÇO") is not -1):
				print "chegou"

				if (not checkAnswer(auxName[auxCounter], auxName, auxText)) and (auxName[auxCounter] != "DEPILACAO.LASER.BR"):
					#click to answer
					elem = browser.find_element_by_xpath("//ul[@class='uiList _4kg  _4kt _6-h _6-j']/div[" + str(auxCounter+1+respostas) + "]/div/div/div[3]/div[1]/a[2]")
					browser.execute_script("arguments[0].click();", elem)
					sleep(1)
					browser.find_element_by_xpath("//div[@class='_4u-c _5n4j']/div[1]/div[2]/div/div/div/div/label/input").send_keys(u"São mais de 370 unidades em todo Brasil. Inscreva-se clicando no link abaixo da imagem, aguarde contato telefônico e aproveite.\n")
					sleep(3)
					respostas = respostas + 1

		except Exception as e:
			print e
			pass
		auxCounter = auxCounter + 1

	return respostas

def initFunction(url, url_inbox):
	#open the google chrome
	browser = webdriver.Chrome()

	#open the whatsapp web
	browser.get(url)

	browser.maximize_window()

	#wait for the login of the user
	wait = raw_input("Aperte qualquer tecla assim que logar no facebook: ")

	#give some time to update the page
	sleep(10)

	#try to go directly to the inbox
	browser.get(url_inbox)

	#give some time to update the page
	sleep(10)

	#click in the instagram part
	browser.find_elements_by_class_name("_6ie2")[2].click()

	sleep(10)

	return browser

def selectTheCampaigns(browser):
	global_counter = len(browser.find_elements_by_xpath("//div[@class='_24tx']/div")) - 2

	total_answered = 0
	preto = True

	for counter in range(global_counter):
		erros = 0
		while True:
			try:
				try:
					if counter == 0:
						auxiliar = browser.find_element_by_xpath("//div[@class='_24tx']/div[" + str(counter + 1) + "]/div[@class='_4k8w _75uw _5_n1 _2tms _5m16']")
						preto = False
					else:
						auxiliar = browser.find_element_by_xpath("//div[@class='_24tx']/div[" + str(counter + 1) + "]/div[@class='_4k8w _75uw _5_n1 _5m16']")
						preto = False

				except:
					if counter == 0:
						print browser.find_element_by_xpath("//div[@class='_24tx']/div[" + str(counter + 1) + "]/div[@class='_4k8w _75uw _5_n1 _2tms _284c _5m16']")
					else:
						print browser.find_element_by_xpath("//div[@class='_24tx']/div[" + str(counter + 1) + "]/div[@class='_4k8w _75uw _5_n1 _284c _5m16']")				
				browser.find_elements_by_xpath("//div[@class='_24tx']/div")[counter].click()
				break
			except Exception as e:
				sleep(1)
				erros = erros + 1
				if erros == 3:
					erros = 0
					counter = counter + 1

				print e
				continue


		sleep(10)

		nome, mensagem = getNameAndMessage(browser)
		answered = answer(browser, nome, mensagem)

		if answered == 0:
			if preto:
				browser.find_element_by_class_name("_3jcz").click()

		total_answered = total_answered + answered

	return total_answered

def dropDown(browser,erro_dropdown):
	print "arrastar"



	if erro_dropdown == 3:
		source_elem = browser.find_element_by_xpath("//div[@class='_1t0r _1t0s _4jdr _1t0v']/div")
		erro_dropdown = 0
	else:
		source_elem = browser.find_element_by_xpath("//div[@class='_1t0r _1t0s _4jdr']/div")				

	source_elem.click()

	sleep(5)

	dest_elem = browser.find_elements_by_xpath("//div[@class='_24tx']/div")[9]

	sleep(5)

	ActionChains(browser).drag_and_drop(source_elem, dest_elem).perform()

def fillSheet(respostas_total):
	book = xlrd.open_workbook("Respondido.xlsx")
	sh = book.sheet_by_index(0)

	rows = sh.nrows

	del book

	wb = openpyxl.load_workbook("Respondido.xlsx")

	ws = wb.get_sheet_by_name("Plan1")

	ws['A' + str(rows + 1)] = str(datetime.now().day) + "/" + str(datetime.now().month)
	ws['B' + str(rows + 1)] = str(respostas_total)

	wb.save("Respondido.xlsx")

def main(erro_dropdown):
	respostas_total = 0

	#the url of the facebook
	url = 'https://www.facebook.com/'

	#the url of the inbox
	url_inbox = 'https://business.facebook.com/depilacao.laser.br/inbox/?business_id=1163510226997646&mailbox_id=275176122991882&selected_item_id=100001503612000'

	#the number of drop-downs
	N = 3

	browser = initFunction(url, url_inbox)

	for counter in range(N):		
		respostas_total = respostas_total + selectTheCampaigns(browser)
		while True:
			try:
				dropDown(browser,erro_dropdown)
				erro_dropdown = 0
				break
			except Exception as e:
				print e
				erro_dropdown = erro_dropdown + 1
				sleep(5)
				continue
		sleep(3)

	fillSheet(respostas_total)

main(erro_dropdown)

#a regular counter
#counter = 0


#END OF THE DROP DOWN


##BELOW THIS SECTION YOU CAN READ THE INSTAGRAM COMMENTS

#print len(browser.find_elements_by_xpath("//ul[@class='uiList _4kg  _4kt _6-h _6-j']/div"))

#while counter < len(browser.find_elements_by_xpath("//div[@class='_24tx']/div")):
	#auxCounter = 0
	#browser.find_elements_by_xpath("//div[@class='_24tx']/div")[counter].click()
	#sleep(5)
	#while auxCounter < len(browser.find_elements_by_xpath("//ul[@class='uiList _4kg  _4kt _6-h _6-j']/div")):
		#try:
		#	print browser.find_element_by_xpath("//ul[@class='uiList _4kg  _4kt _6-h _6-j']/div[" + str(auxCounter+1) + "]").text
	#	except:
		#	pass
		#auxCounter = auxCounter + 1
	#counter = counter + 1

#print div_types

#END OF THE SECTION OF READING INSTAGRAM COMMENTS