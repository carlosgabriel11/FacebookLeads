#!/usr/bin/env python
# -*- coding: utf-8 -*-

import requests
import xlrd
import openpyxl
from time import sleep
from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime

#a lista que irá conter as contas de anúncio
adAccounts = []

#conta 1
adAccounts.append('1906220526059942')
#conta 2
adAccounts.append('111100286392685')
#conta 3
adAccounts.append('101373384038786')
#conta 4
adAccounts.append('1971824669499527')
#conta 5
adAccounts.append('2068943549787638')
#conta 6
adAccounts.append('101964773538755')
#conta 7
adAccounts.append('294354761190803')
#conta 8
adAccounts.append('1249262278555646')
#conta 9
adAccounts.append('403669640487196')

#nome da planilha
nome_planilha = 'C:\\Users\\renderxp\\Dropbox\\_Relatorios\\_Consolidados\\Gastos.xlsx'
#nome indicador
nome_indicador = 'C:\\Users\\renderxp\\Dropbox\\_Relatorios\\_Consolidados\\Indicadores.xlsx'

#função para pegar o mês de ontem
def getMonthYesterday():
	#se está no primeiro dia do mês, tem que ser referente ao mês anterior
	if datetime.now().day == 1:
		#se o mês anterior for menor que 10, coloca-se o 0 na frente
		if (datetime.now().month - 1) < 10:
			return '0' + str(datetime.now().month - 1)
		else:
			return str(datetime.now().month - 1)

	else:
		#se o mês for menor que 10, coloca-se o 0 na frente
		if datetime.now().month < 10:
			return '0' + str(datetime.now().month)
		else:
			return str(datetime.now().month)

#função para pegar o mês anterior
def getMonthBefore(num):
	#se o dia atual menos a quantidade desejada for menor ou igual a 0, leva o mês anterior
	if (datetime.now().day - num) <= 0:
		mes = datetime.now().month - 1
		#quanto ainda falta tirar para o mês anterior
		dia = num - datetime.now().day

		#loop para pegar o mês exato
		while True:
			if (mes == 1) or (mes == 3) or (mes == 5) or (mes == 7) or (mes == 8) or (mes == 10):
				if (31 - dia) <= 0:
					dia -= 31
					mes -= 1
					continue
				else:
					#se o mês for menor que 10, coloca-se o 0 na frente
					if mes < 10:
						return '0' + str(mes)
					else:
						return str(mes)

			elif (mes == 4) or (mes == 6) or (mes == 9) or (mes == 11):
				if (30 - dia) <= 0:
					dia -= 30
					mes -= 1
					continue
				else:
					#se o mês for menor que 10, coloca-se o 0 na frente
					if mes < 10:
						return '0' + str(mes)
					else:
						return str(mes)

			else:
				if (28 - dia) <= 0:
					dia -= 28
					mes -= 1
					continue
				else:
					#se o mês for menor que 10, coloca-se o 0 na frente
					if mes < 10:
						return '0' + str(mes)
					else:
						return str(mes)

	else:
		#se o mês for menor que 10, coloca-se o 0 na frente
		if datetime.now().month < 10:
			return '0' + str(datetime.now().month)
		else:
			return str(datetime.now().month)

#função para pegar o mês de hoje
def getMonthToday():
	#se o mês atual for menor que 10, coloca-se o 0 na frente
	if datetime.now().month < 10:
		return '0' + str(datetime.now().month)


	else:
		return str(datetime.now().month)

#função para pegar o dia de ontem
def getDayYesterday():
	#se está no primeiro dia do mês, tem que pegar o último dia do mês anterior
	if datetime.now().day == 1:
		if ((datetime.now().month - 1) == 4) or ((datetime.now().month - 1) == 6) or ((datetime.now().month - 1) == 9) or ((datetime.now().month - 1) == 11):
			return str(30)

		else:
			return str(31) 

	else:
		#se o dia de ontem for menor que 10, coloca-se o 0 na frente
		if (datetime.now().day - 1) < 10:
			return '0' + str(datetime.now().day - 1)
		else:
			return str(datetime.now().day - 1)

#função para pegar o dia anterior
def getDayBefore(num):
	#se o dia atual menos a quantidade de dias desejadas
	if (datetime.now().day - num) <= 0:
		mes = datetime.now().month - 1
		#quanto ainda falta tirar para o mês anterior
		dia = num - datetime.now().day

		#loop para pegar o mês exato
		while True:
			if (mes == 1) or (mes == 3) or (mes == 5) or (mes == 7) or (mes == 8) or (mes == 10):
				if (31 - dia) <= 0:
					dia -= 31
					mes -= 1
					continue
				else:
					dia = 31 - dia
					#se o dia for menor que 10, coloca-se o 0 na frente
					if dia < 10:
						return '0' + str(dia)
					else:
						return str(dia)

			elif (mes == 4) or (mes == 6) or (mes == 9) or (mes == 11):
				if (30 - dia) <= 0:
					dia -= 30
					mes -= 1
					continue
				else:
					dia = 30 - dia
					#se o dia for menor que 10, coloca-se o 0 na frente
					if dia < 10:
						return '0' + str(dia)
					else:
						return str(dia)

			else:
				if (28 - dia) <= 0:
					dia -= 28
					mes -= 1
					continue
				else:
					dia = 28 - dia
					#se o dia for menor que 10, coloca-se o 0 na frente
					if dia < 10:
						return '0' + str(dia)
					else:
						return str(dia) 

	else:
		#se o dia de ontem for menor que 10, coloca-se o 0 na frente
		if (datetime.now().day - num) < 10:
			return '0' + str(datetime.now().day - num)
		else:
			return str(datetime.now().day - num)

#função para pegar o mês de hoje
def getDayToday():
	#se o dia de hoje for menor que 10, coloca-se o 0 na frente
	if datetime.now().day < 10:
		return '0' + str(datetime.now().day)

	else:
		return str(datetime.now().day)

def getVariation(number):
	book = xlrd.open_workbook(nome_indicador)


	linha_aproveitamento = 3

	coluna = 0

	if number == 1:
		sh = book.sheet_by_index(int(datetime.now().month) - 1)
		for counter in range(31):
			if sh.cell_value(0, counter) == int(datetime.now().day):
				coluna = counter

	else:
		sh = book.sheet_by_index(int(getMonthBefore(number - 1)))
		for counter in range(31):
			if sh.cell_value(0, counter) == int(getDayBefore(number - 1)):
				coluna = counter

	while True:
		if sh.cell_value(linha_aproveitamento, coluna) == "0":
			coluna += 1
			continue

		print coluna
		print sh.cell_value(linha_aproveitamento, coluna)
		return float(sh.cell_value(linha_aproveitamento, coluna))

def putSheet(number, cadastros, valores, precos):
	#abrindo a planilha
	book = xlrd.open_workbook(nome_planilha)
	#abrindo aba
	sh = book.sheet_by_index(int(getMonthBefore(number))-1)

	#um contador qualquer
	counter = 2

	#pegar a linha que se deve colocar
	while True:
		if sh.cell_value(rowx = counter, colx = 0) == "":
			rows = counter
			break
		counter += 1

	#apagar o objeto planilha de leitura
	del book

	#abrir a planilha no modo de escrita
	wb = openpyxl.load_workbook(nome_planilha)

	#abrir a planilha
	ws = wb.get_sheet_by_name(wb.get_sheet_names()[int(getMonthBefore(number)) - 1])

	variacao = getVariation(number)

	if number == 1:
		#colocar as informações da primeira conta
		ws['A' + str(rows + 1)] = getDayYesterday() + "/" + getMonthYesterday()
		if cadastros[0].find('.') is not -1:
			cadastros[0] = cadastros[0].split('.')[0] + cadastros[0].split('.')[1]
		ws['B' + str(rows + 1)] = int(cadastros[0])
		valores[0] = valores[0].split(',')[0] + '.' + valores[0].split(',')[1]
		ws['C' + str(rows + 1)] = float(valores[0])

		#colocar as informações da segunda conta
		ws['F' + str(rows + 1)] = getDayYesterday() + "/" + getMonthYesterday()
		if cadastros[1].find('.') is not -1:
			cadastros[1] = cadastros[1].split('.')[0] + cadastros[1].split('.')[1]
		ws['G' + str(rows + 1)] = int(cadastros[1])
		valores[1] = valores[1].split(',')[0] + '.' + valores[1].split(',')[1]
		ws['H' + str(rows + 1)] = float(valores[1])

		#colocar as informações da terceira conta
		ws['K' + str(rows + 1)] = getDayYesterday() + "/" + getMonthYesterday()
		if cadastros[2].find('.') is not -1:
			cadastros[2] = cadastros[2].split('.')[0] + cadastros[2].split('.')[1]
		ws['L' + str(rows + 1)] = int(cadastros[2])
		valores[2] = valores[2].split(',')[0] + '.' + valores[2].split(',')[1]
		ws['M' + str(rows + 1)] = float(valores[2])

		#colocar as informações da quarta conta
		ws['P' + str(rows + 1)] = getDayYesterday() + "/" + getMonthYesterday()
		if cadastros[3].find('.') is not -1:
			cadastros[3] = cadastros[3].split('.')[0] + cadastros[3].split('.')[1]
		ws['Q' + str(rows + 1)] = int(cadastros[3])
		valores[3] = valores[3].split(',')[0] + '.' + valores[3].split(',')[1]
		ws['R' + str(rows + 1)] = float(valores[3])

		#colocar as informações da quinta conta
		ws['U' + str(rows + 1)] = getDayYesterday() + "/" + getMonthYesterday()
		if cadastros[4].find('.') is not -1:
			cadastros[4] = cadastros[4].split('.')[0] + cadastros[4].split('.')[1]
		ws['V' + str(rows + 1)] = int(cadastros[4])
		valores[4] = valores[4].split(',')[0] + '.' + valores[4].split(',')[1]
		ws['W' + str(rows + 1)] = float(valores[4])

		#colocar as informações da sexta conta
		ws['Z' + str(rows + 1)] = getDayYesterday() + "/" + getMonthYesterday()
		if cadastros[5].find('.') is not -1:
			cadastros[5] = cadastros[5].split('.')[0] + cadastros[5].split('.')[1]
		ws['AA' + str(rows + 1)] = int(cadastros[5])
		valores[5] = valores[5].split(',')[0] + '.' + valores[5].split(',')[1]
		ws['AB' + str(rows + 1)] = float(valores[5])

		#colocar as informações da setima conta
		ws['AE' + str(rows + 1)] = getDayYesterday() + "/" + getMonthYesterday()
		if cadastros[6].find('.') is not -1:
			cadastros[6] = cadastros[6].split('.')[0] + cadastros[6].split('.')[1]
		ws['AF' + str(rows + 1)] = int(cadastros[6])
		valores[6] = valores[6].split(',')[0] + '.' + valores[6].split(',')[1]
		ws['AG' + str(rows + 1)] = float(valores[6])

		#colocar as informações da oitava conta
		ws['AJ' + str(rows + 1)] = getDayYesterday() + "/" + getMonthYesterday()
		if cadastros[7].find('.') is not -1:
			cadastros[7] = cadastros[7].split('.')[0] + cadastros[7].split('.')[1]
		ws['AK' + str(rows + 1)] = int(cadastros[7])
		valores[7] = valores[7].split(',')[0] + '.' + valores[7].split(',')[1]
		ws['AL' + str(rows + 1)] = float(valores[7])

		if precos[0].find('.') is not -1:
			precos[0] = precos[0].split('.')[0] + precos[0].split('.')[1]
		precos[0] = precos[0].split(',')[0] + '.' + precos[0].split(',')[1]
		ws['D' + str(rows + 1)] = float(precos[0])

		if precos[1].find('.') is not -1:
			precos[1] = precos[1].split('.')[0] + precos[1].split('.')[1]
		precos[1] = precos[1].split(',')[0] + '.' + precos[1].split(',')[1]
		ws['I' + str(rows + 1)] = float(precos[1])

		if precos[2].find('.') is not -1:
			precos[2] = precos[2].split('.')[0] + precos[2].split('.')[1]
		precos[2] = precos[2].split(',')[0] + '.' + precos[2].split(',')[1]
		ws['N' + str(rows + 1)] = float(precos[2])

		if precos[3].find('.') is not -1:
			precos[3] = precos[3].split('.')[0] + precos[3].split('.')[1]
		precos[3] = precos[3].split(',')[0] + '.' + precos[3].split(',')[1]
		ws['S' + str(rows + 1)] = float(precos[3])

		if precos[4].find('.') is not -1:
			precos[4] = precos[4].split('.')[0] + precos[4].split('.')[1]
		precos[4] = precos[4].split(',')[0] + '.' + precos[4].split(',')[1]
		ws['X' + str(rows + 1)] = float(precos[4])

		if precos[5].find('.') is not -1:
			precos[5] = precos[5].split('.')[0] + precos[5].split('.')[1]
		precos[5] = precos[5].split(',')[0] + '.' + precos[5].split(',')[1]
		ws['AC' + str(rows + 1)] = float(precos[5])

		if precos[6].find('.') is not -1:
			precos[6] = precos[6].split('.')[0] + precos[6].split('.')[1]
		precos[6] = precos[6].split(',')[0] + '.' + precos[6].split(',')[1]
		ws['AH' + str(rows + 1)] = float(precos[6])		

		if precos[7].find('.') is not -1:
			precos[7] = precos[7].split('.')[0] + precos[7].split('.')[1]
		precos[7] = precos[7].split(',')[0] + '.' + precos[7].split(',')[1]
		ws['AM' + str(rows + 1)] = float(precos[7])

		ws['AR' + str(rows + 1)] = variacao

		wb.save(nome_planilha)

	else:
		#colocar as informações da primeira conta
		ws['A' + str(rows + 1)] = getDayBefore(number) + "/" + getMonthBefore(number)
		if cadastros[0].find('.') is not -1:
			cadastros[0] = cadastros[0].split('.')[0] + cadastros[0].split('.')[1]
		print cadastros[0]
		ws['B' + str(rows + 1)] = int(cadastros[0])
		valores[0] = valores[0].split(',')[0] + '.' + valores[0].split(',')[1]
		ws['C' + str(rows + 1)] = float(valores[0])

		#colocar as informações da segunda conta
		ws['F' + str(rows + 1)] = getDayBefore(number) + "/" + getMonthBefore(number)
		if cadastros[1].find('.') is not -1:
			cadastros[1] = cadastros[1].split('.')[0] + cadastros[1].split('.')[1]
		ws['G' + str(rows + 1)] = int(cadastros[1])
		valores[1] = valores[1].split(',')[0] + '.' + valores[1].split(',')[1]
		ws['H' + str(rows + 1)] = float(valores[1])

		#colocar as informações da terceira conta
		ws['K' + str(rows + 1)] = getDayBefore(number) + "/" + getMonthBefore(number)
		if cadastros[2].find('.') is not -1:
			cadastros[2] = cadastros[2].split('.')[0] + cadastros[2].split('.')[1]
		ws['L' + str(rows + 1)] = int(cadastros[2])
		valores[2] = valores[2].split(',')[0] + '.' + valores[2].split(',')[1]
		ws['M' + str(rows + 1)] = float(valores[2])

		#colocar as informações da quarta conta
		ws['P' + str(rows + 1)] = getDayBefore(number) + "/" + getMonthBefore(number)
		if cadastros[3].find('.') is not -1:
			cadastros[3] = cadastros[3].split('.')[0] + cadastros[3].split('.')[1]
		ws['Q' + str(rows + 1)] = int(cadastros[3])
		valores[3] = valores[3].split(',')[0] + '.' + valores[3].split(',')[1]
		ws['R' + str(rows + 1)] = float(valores[3])

		#colocar as informações da quinta conta
		ws['U' + str(rows + 1)] = getDayBefore(number) + "/" + getMonthBefore(number)
		if cadastros[4].find('.') is not -1:
			cadastros[4] = cadastros[4].split('.')[0] + cadastros[4].split('.')[1]
		ws['V' + str(rows + 1)] = int(cadastros[4])
		valores[4] = valores[4].split(',')[0] + '.' + valores[4].split(',')[1]
		ws['W' + str(rows + 1)] = float(valores[4])

		#colocar as informações da sexta conta
		ws['Z' + str(rows + 1)] = getDayBefore(number) + "/" + getMonthBefore(number)
		if cadastros[5].find('.') is not -1:
			cadastros[5] = cadastros[5].split('.')[0] + cadastros[5].split('.')[1]
		ws['AA' + str(rows + 1)] = int(cadastros[5])
		valores[5] = valores[5].split(',')[0] + '.' + valores[5].split(',')[1]
		ws['AB' + str(rows + 1)] = float(valores[5])

		#colocar as informações da setima conta
		ws['AE' + str(rows + 1)] = getDayBefore(number) + "/" + getMonthBefore(number)
		if cadastros[6].find('.') is not -1:
			cadastros[6] = cadastros[6].split('.')[0] + cadastros[6].split('.')[1]
		ws['AF' + str(rows + 1)] = int(cadastros[6])
		valores[6] = valores[6].split(',')[0] + '.' + valores[6].split(',')[1]
		ws['AG' + str(rows + 1)] = float(valores[6])

		#colocar as informações da oitava conta
		ws['AJ' + str(rows + 1)] = getDayBefore(number) + "/" + getMonthBefore(number)
		if cadastros[7].find('.') is not -1:
			cadastros[7] = cadastros[7].split('.')[0] + cadastros[7].split('.')[1]
		ws['AK' + str(rows + 1)] = int(cadastros[7])
		valores[7] = valores[7].split(',')[0] + '.' + valores[7].split(',')[1]
		ws['AL' + str(rows + 1)] = float(valores[7])


		if precos[0].find('.') is not -1:
			precos[0] = precos[0].split('.')[0] + precos[0].split('.')[1]
		precos[0] = precos[0].split(',')[0] + '.' + precos[0].split(',')[1]
		ws['D' + str(rows + 1)] = float(precos[0])

		if precos[1].find('.') is not -1:
			precos[1] = precos[1].split('.')[0] + precos[1].split('.')[1]
		precos[1] = precos[1].split(',')[0] + '.' + precos[1].split(',')[1]
		ws['I' + str(rows + 1)] = float(precos[1])

		if precos[2].find('.') is not -1:
			precos[2] = precos[2].split('.')[0] + precos[2].split('.')[1]
		precos[2] = precos[2].split(',')[0] + '.' + precos[2].split(',')[1]
		ws['N' + str(rows + 1)] = float(precos[2])

		if precos[3].find('.') is not -1:
			precos[3] = precos[3].split('.')[0] + precos[3].split('.')[1]
		precos[3] = precos[3].split(',')[0] + '.' + precos[3].split(',')[1]
		ws['S' + str(rows + 1)] = float(precos[3])

		if precos[4].find('.') is not -1:
			precos[4] = precos[4].split('.')[0] + precos[4].split('.')[1]
		precos[4] = precos[4].split(',')[0] + '.' + precos[4].split(',')[1]
		ws['X' + str(rows + 1)] = float(precos[4])

		if precos[5].find('.') is not -1:
			precos[5] = precos[5].split('.')[0] + precos[5].split('.')[1]
		precos[5] = precos[5].split(',')[0] + '.' + precos[5].split(',')[1]
		ws['AC' + str(rows + 1)] = float(precos[5])

		if precos[6].find('.') is not -1:
			precos[6] = precos[6].split('.')[0] + precos[6].split('.')[1]
		precos[6] = precos[6].split(',')[0] + '.' + precos[6].split(',')[1]
		ws['AH' + str(rows + 1)] = float(precos[6])		

		if precos[7].find('.') is not -1:
			precos[7] = precos[7].split('.')[0] + precos[7].split('.')[1]
		precos[7] = precos[7].split(',')[0] + '.' + precos[7].split(',')[1]
		ws['AM' + str(rows + 1)] = float(precos[7])

		ws['AR' + str(rows + 1)] = variacao

		wb.save(nome_planilha)
				

#função para iniciar o programa
def init(url, url_adaccount, number):
	#o numero de cadastros em cada conta
	cadastros = []
	#os valores em cada conta
	valores = []
	#os precos de cada conta
	precos = []

	#abrir o google chrome
	browser = webdriver.Chrome()

	#abrir o facebook
	browser.get(url)

	#maximizar a tela
	browser.maximize_window()

	#esperar o usuario fazer login
	wait = raw_input("Aperte qualquer tecla assim que logar no facebook: ")

	#tempo para a página atualizar
	sleep(10)

	if int(number) == 1:
		#ir em cada conta de anuncio
		for conta in adAccounts:
			try:
				nova_url = url_adaccount + conta + '&business_id=1163510226997646&date=2019-' + getMonthYesterday() + '-' + getDayYesterday() + '_2019-' + getMonthToday() + '-' + getDayToday()			
				nova_url = nova_url + '%2Cyesterday&comparision_date='

				#abrir a url para cada conta de anúncio
				browser.get(nova_url)

				#tempo de espera
				sleep(15)

				#para o caso específico da conta 6
				if conta != '101964773538755':
					cadastros.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[3]/div/div/div/div[1]").text)
					valores.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[6]/div/div/div/div[1]/span").text.split()[1])
					precos.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[7]/div/div/div/div[1]/span").text.split()[1])
				else:
					browser.find_element_by_xpath("//a[@class=' _5bbf _3-9a _55pi _2agf _4o_4 _4jy0 _4jy4 _517h _51sy _42ft']").click()
					sleep(7)
					browser.find_element_by_xpath("//div[@class='uiContextualLayer uiContextualLayerBelowLeft']/div/div/div/div/div[6]/div").click()
					sleep(7)
					browser.find_element_by_xpath("//div[@class='uiContextualLayer uiContextualLayerBelowLeft']/div/div/div/div/div[6]/div/ul/li[7]/div").click()
					sleep(7)
					cadastros.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[3]/div/div/div/div[1]").text)
					valores.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[6]/div/div/div/div[1]/span").text.split()[1])
					precos.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[7]/div/div/div/div[1]/span").text.split()[1])

			except:

				cadastros[len(cadastros) - 1] = '0'
				valores.append(u'0,0')
				precos.append(u'0,0')
				continue
		putSheet(1, cadastros, valores, precos)

		cadastros = []
		valores = []
		precos = []
	else:
		contador = int(number)
		while contador > 0:
			if contador != 1:
				#ir em cada conta de anuncio
				for conta in adAccounts:
					try:
						nova_url = url_adaccount + conta + '&business_id=1163510226997646&date=2019-' + getMonthBefore(contador) + '-' + getDayBefore(contador) + '_2019-' + getMonthBefore(contador - 1) + '-' + getDayBefore(contador - 1)			
						nova_url = nova_url + '&comparision_date='

						#abrir a url para cada conta de anúncio
						browser.get(nova_url)

						#tempo de espera
						sleep(15)

						#para o caso específico da conta 6
						if conta != '101964773538755':
							cadastros.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[3]/div/div/div/div[1]").text)
							valores.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[6]/div/div/div/div[1]/span").text.split()[1])
							precos.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[7]/div/div/div/div[1]/span").text.split()[1])
						else:
							browser.find_element_by_xpath("//a[@class=' _5bbf _3-9a _55pi _2agf _4o_4 _4jy0 _4jy4 _517h _51sy _42ft']").click()
							sleep(7)
							browser.find_element_by_xpath("//div[@class='uiContextualLayer uiContextualLayerBelowLeft']/div/div/div/div/div[6]/div").click()
							sleep(7)
							browser.find_element_by_xpath("//div[@class='uiContextualLayer uiContextualLayerBelowLeft']/div/div/div/div/div[6]/div/ul/li[7]/div").click()
							sleep(7)
							cadastros.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[3]/div/div/div/div[1]").text)
							valores.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[6]/div/div/div/div[1]/span").text.split()[1])
							precos.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[7]/div/div/div/div[1]/span").text.split()[1])
					except:
						cadastros[len(cadastros) - 1] = '0'
						valores.append(u'0,0')
						precos.append(u'0,0')
						continue

				putSheet(contador, cadastros, valores, precos)

				cadastros = []
				valores = []
				precos = []
			else:
				#ir em cada conta de anuncio
				for conta in adAccounts:
					try:
						nova_url = url_adaccount + conta + '&business_id=1163510226997646&date=2019-' + getMonthYesterday() + '-' + getDayYesterday() + '_2019-' + getMonthToday() + '-' + getDayToday()			
						nova_url = nova_url + '%2Cyesterday&comparision_date='

						#abrir a url para cada conta de anúncio
						browser.get(nova_url)

						#tempo de espera
						sleep(15)

						#para o caso específico da conta 6
						if conta != '101964773538755':
							cadastros.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[3]/div/div/div/div[1]").text)
							valores.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[6]/div/div/div/div[1]/span").text.split()[1])
							precos.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[7]/div/div/div/div[1]/span").text.split()[1])
						else:
							browser.find_element_by_xpath("//a[@class=' _5bbf _3-9a _55pi _2agf _4o_4 _4jy0 _4jy4 _517h _51sy _42ft']").click()
							sleep(7)
							browser.find_element_by_xpath("//div[@class='uiContextualLayer uiContextualLayerBelowLeft']/div/div/div/div/div[6]/div").click()
							sleep(7)
							browser.find_element_by_xpath("//div[@class='uiContextualLayer uiContextualLayerBelowLeft']/div/div/div/div/div[6]/div/ul/li[7]/div").click()
							sleep(7)
							cadastros.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[3]/div/div/div/div[1]").text)
							valores.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[6]/div/div/div/div[1]/span").text.split()[1])
							precos.append(browser.find_element_by_xpath("//div[@class='_5d6f']/div/div[1]/div[4]/div/div/div[2]/div/div[7]/div/div/div/div[1]/span").text.split()[1])

					except:
						cadastros[len(cadastros) - 1] = '0'
						valores.append(u'0,0')
						precos.append(u'0,0')
						continue

				putSheet(1, cadastros, valores, precos)
				cadastros = []
				valores = []
				precos = []

			contador -= 1



def main():
	#the url of facebook
	url = 'https://www.facebook.com/'

	#the url of the inbox
	url_adaccount = 'https://business.facebook.com/adsmanager/manage/campaigns?act='

	number = raw_input("Digite a quantidade de dias que gostaria de pegar: ")

	#print getMonthBefore(int(number))
	#print getDayBefore(int(number))

	init(url, url_adaccount, number)

main()